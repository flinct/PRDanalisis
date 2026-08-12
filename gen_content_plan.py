"""Generate 30-day LDR couple content plan .xlsx. Data from riset TikTok ID trend 2026."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

rows = [
    # day, host, judul, format/tren, hook, caption+CTA, sound
    (1,"A","Day 1 — Pagi di kota A","POV","Teks besar: 'POV: pacarmu 500km jauhnya'","Mulai 30 hari kami. Kalian tim LDR gak?","sound lokal trending"),
    (2,"B","Day 2 — Malam di kota B","POV sambung","'Sementara dia bangun, aku baru mau tidur'","Beda 1 jam tapi rasa kangen sama. #pasanganjarakjauh","sound sama day 1"),
    (3,"A","GRWM buat video call sama dia","GRWM","'Get ready... buat nelpon doi 3 jam'","GRWM versi LDR. Kalian dandan gak pas VC?","GRWM trending sound"),
    (4,"B","Reaksi dia waktu bilang kangen","Screen-record chat","'Aku screenshot chat ini...'","Reaksi random tapi bikin meleleh. Pernah gini?","voice note asli"),
    (5,"A","25 hari lagi ketemu","Countdown","Angka besar '25'","Countdown dimulai! Kalian LDR berapa lama sekali ketemu?","emosional/piano"),
    (6,"B","Unspoken rules LDR kami","List carousel","'6 aturan gak tertulis LDR kami'","Rule ke-4 paling penting. Kalian punya rule apa?","upbeat"),
    (7,"A","Bird Test versi LDR","Trend global","'Aku tunjukin burung ke dia via VC...'","Bird test tapi jarak jauh. Dia lulus gak ya?","trend bird test"),
    (8,"B","Ritual 5-9 pas LDR","5-9 ritual","'5 sore sampai 9 malam versi LDR'","Dinner call + journaling. Ritual malam kalian apa?","cozy/lofi"),
    (9,"A","Date night jarak jauh: dinner call","POV","'Dinner tapi beda kota'","Nonton film bareng lewat layar. Kalian date jarak jauh gimana?","romantic"),
    (10,"B","Tantangan: pesan makanan buat dia","Challenge","'Aku pesenin makanan ke kota dia'","Surprise GoFood lintas kota. Berhasil gak?","fun/suspense"),
    (11,"A","Topik wajib tiap telepon malem","List","'5 hal yang selalu kami omongin'","Topik No.3 gak pernah abis. Kalian bahas apa tiap malem?","chill"),
    (12,"B","Kota B vs kota A","Vote/duet","'Kota mana lebih enak?'","Vote di komen! Kota kalian menang gak?","upbeat"),
    (13,"A","Kode rahasia kami di chat","Screen-record","'Arti emoji ini cuma kami yang tau'","Kode rahasia pasangan. Kalian punya bahasa sendiri?","playful"),
    (14,"B","Reaksi voice note random","Reaction","'Dia kirim VN jam 2 pagi...'","Isi VN-nya bikin ngakak. Pernah dapet VN random?","voice note asli"),
    (15,"A","AI generate masa depan kita","Trend global","'Aku generate foto masa depan kami pakai AI'","Hasilnya... mirip gak ya? #ldrindonesia","trend AI future"),
    (16,"B","GRWM malam minggu sendirian","GRWM","'GRWM tapi malmingnya sama HP'","Malming LDR = VC-an. Kalian malming gimana?","GRWM sound"),
    (17,"A","Hal yang aku beli biar makin deket","#TikTokMadeMeBuyIt","'3 barang wajib anak LDR'","Matching bracelet + lampu sync. Kalian punya?","review sound"),
    (18,"B","Hopecore: kenapa LDR worth it","#hopecore","'Buat yang mau nyerah sama LDR...'","Jarak cuma angka. Kalian kuat! Setuju?","soft/uplifting"),
    (19,"A","Malam minggu sendirian versi LDR","POV","'Malming tapi dia di kota lain'","Sepi tapi worth it. Relate gak?","melankolis"),
    (20,"B","Cara kami ngatasi salah paham","Tips+story","'Kami hampir putus karena chat...'","Cara selesain berantem jarak jauh. Tips kalian apa?","serius/lembut"),
    (21,"A","9 hari lagi","Countdown","Angka besar '9'","Tinggal seminggu lebih! Excited gak liat progress ini?","building tension"),
    (22,"B","GRWM buat ketemu besok","GRWM","'GRWM H-1 ketemu doi'","Nervous tapi happy. Kalian gimana pas mau ketemu?","GRWM excited sound"),
    (23,"A","Playlist perjalanan buat dia","Share","'Lagu yang aku dengerin di jalan ke dia'","Playlist LDR kami. Rekomen lagu dong di komen!","lagu playlist"),
    (24,"B","5-9 minggu terakhir: gak fokus","5-9 ritual","'Ritual sore tapi pikiran ke dia terus'","Gak bisa fokus kerja. Kalian gitu juga gak?","cozy"),
    (25,"A","Surprise yang aku siapin","Hook/teaser","'Aku siapin sesuatu buat dia...' (blur)","Apa ya kira-kira? Tebak di komen!","suspense"),
    (26,"B","T-3: nervous","Vlog singkat","'3 hari lagi... deg-degan'","Nervous parah. Normal gak sih?","heartbeat sound"),
    (27,"A","T-1: malam sebelum ketemu","POV","'Malam terakhir sebelum ketemu'","Gak bisa tidur. Kalian gitu juga?","emosional"),
    (28,"B","HARI INI KETEMU!","POV meeting day","'Detik-detik ketemu setelah sebulan'","MOMEN INI!! Nangis gak nangis? #pasanganjarakjauh","klimaks/emosional"),
    (29,"A","Recap hari ketemu","Montage","'Rangkuman hari paling ditunggu'","Semua worth it. Fav moment kalian yang mana?","montage emosional"),
    (30,"B","Day 30 — pulang + teaser","Ending hook","'Pulang lagi... tapi bulan depan?'","Loop lagi bulan depan! Follow biar gak ketinggalan.","bittersweet + hook"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "30 Hari Content Plan"

headers = ["Day","Host","Judul Konten","Format / Tren","Hook (3 detik)","Caption + CTA","Sound/Audio"]
ws.append(headers)

# style header
hfill = PatternFill("solid", fgColor="C00000")
hfont = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
for c in ws[1]:
    c.fill=hfill; c.font=hfont; c.border=border
    c.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)

# rows, alternate host color
fillA = PatternFill("solid", fgColor="FCE4EC")  # pink = host A
fillB = PatternFill("solid", fgColor="E3F2FD")  # blue = host B
for r in rows:
    ws.append(r)
    fill = fillA if r[1]=="A" else fillB
    for c in ws[ws.max_row]:
        c.fill=fill; c.border=border
        c.alignment=Alignment(vertical="top", wrap_text=True)

widths=[6,7,32,18,32,42,22]
for i,w in enumerate(widths,1):
    ws.column_dimensions[chr(64+i)].width=w
ws.freeze_panes="A2"

# sheet 2: strategi & hashtag
ws2 = wb.create_sheet("Strategi & Hashtag")
strat = [
    ("KATEGORI","DETAIL"),
    ("Konsep","'30 Hari 1 Cerita' — post nyambung tiap hari, bergantian Host A (ganjil) & Host B (genap), berakhir di hari ketemu lalu loop"),
    ("Niche target","LDR Indonesia — pasar kosong, hashtag lokal masih kecil = peluang besar"),
    ("Hashtag utama","#ldrindonesia #pasanganjarakjauh #ldr #longdistancerelationship #coupleindonesia"),
    ("Hashtag global (views/post tinggi)","#longdistancecouple (28.596 avg views/post - niche padat) #longdistancelove"),
    ("Skip","#fyp — kebanjiran, gak efektif buat niche kecil"),
    ("Tren lokal dominan 2026","GRWM, POV, 5-9 ritual, #hopecore, #TikTokMadeMeBuyIt (sumber: IDN Times)"),
    ("Tren global (porsi kecil)","Bird Test, Strawberry Test, AI future family — taruh minggu awal buat momentum"),
    ("Jadwal post","1 post/hari + 3 story (pagi/siang/malam). Prime time 19.00-21.00 WIB, Selasa-Jumat"),
    ("Konsistensi feed","Cover format sama tiap post: 'Day X' + template = keliatan series"),
    ("Sound","Trending sound lokal — cek lagitren.id/tiktok tiap hari + 1 voice note asli kalian buat ciri khas"),
    ("Growth","Balas semua komen, duet konten LDR lain, repost story follower, caption selalu tanya/poll"),
    ("Seasonal","Interest LDR naik Nov-Des (gifting). 30 hari sekarang = bangun audiens sebelum peak"),
    ("Riset harian","Cek trending sound + adapt 1 ke jadwal tiap hari"),
]
for r in strat: ws2.append(r)
for c in ws2[1]:
    c.fill=hfill; c.font=hfont; c.border=border
    c.alignment=Alignment(horizontal="center", vertical="center")
for row in ws2.iter_rows(min_row=2):
    for c in row:
        c.border=border; c.alignment=Alignment(vertical="top", wrap_text=True)
    row[0].font=Font(bold=True)
ws2.column_dimensions["A"].width=32
ws2.column_dimensions["B"].width=80
ws2.freeze_panes="A2"

out = r"C:\Users\MyBook SAGA 12\Desktop\PRDanalisis\Content_Plan_LDR_30Hari.xlsx"
wb.save(out)

# self-check
wb2 = openpyxl.load_workbook(out)
assert wb2["30 Hari Content Plan"].max_row == 31, "harus 30 hari + header"
assert wb2.sheetnames == ["30 Hari Content Plan","Strategi & Hashtag"]
print("OK ->", out)

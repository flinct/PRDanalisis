from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment

SRC = Path(__file__).with_name('satuinbox_conversation.messages.training.json')
OUT = Path(__file__).with_name('sample_100.xlsx')


def main():
    data = json.loads(SRC.read_text(encoding='utf-8'))
    wb = Workbook()
    ws = wb.active
    ws.title = 'sample_100'
    ws.append(['conversationId', 'inbound', 'outbound'])

    for cid, msgs in list(data.items())[:100]:
        inbound = '\n'.join(m['value'] for m in msgs if m['type'] == 'inbound')
        outbound = '\n'.join(m['value'] for m in msgs if m['type'] == 'outbound')
        ws.append([cid, inbound, outbound])

    for col in ('A', 'B', 'C'):
        for cell in ws[col]:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 80

    wb.save(OUT)
    print(OUT)
    print('rows:', ws.max_row - 1)
    assert ws.max_row == 101, ws.max_row


if __name__ == '__main__':
    main()
